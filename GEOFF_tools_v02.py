__version__ = "v0.2.2"  #major,minor,patch
#date 2024_09_24

import sys
import argparse
import os
from collections import defaultdict
import openpyxl
import yaml
from yaml.loader import   Reader
import re
import csv

##TODO:  code V:url   and V:http:// url  and
##TODO:  add file and directory path confirmation
##TODO:  add a look up for gps
##fixed settings that might change --but not by users but by programmer
para={
    "tsv": ["study","site","prev","validate"],
    "file_suffix_study": "_study_data",
    "file_suffix_site": "_site_data",
    "file_suffix_prev": "_prevalence_data",
    "file_suffix_validate":"_validate_parameters",
    "log_file": None,
    "log_error_file":None,
    "iso3c_africa": [
        "DZA", "AGO", "BEN", "BWA", "BFA", "BDI", "CMR", "CPV", "CAF", "TCD", "COM", "COG",
        "COD", "CIV", "DJI", "EGY", "GNQ", "ERI", "ETH", "GAB", "GMB", "GHA", "GIN", "GNB", "KEN", "LSO",
        "LBR", "LBY", "MDG", "MLI", "MWI", "MRT", "MUS", "MYT", "MAR", "MOZ", "NAM", "NER", "NGA",
        "REU", "RWA", "STP", "SEN", "SYC", "SLE", "SOM", "ZAF",
        "SSD", "SDN", "SWZ", "TZA", "TGO", "TUN", "UGA", "ESH", "ZMB", "ZWE",
    ],
    "prev_field_format": {
        'site_uid':{'F': ['lowercase'], 'P': [ r'^[\w-]+$' ] , 'R':['yes']},
        'substudy':{'F': ['lowercase'], 'P': [ r'^^[\w-]+$' ] },
        'date_start':{'F': ['date'], 'V':['datepartial'], 'R':['yes']},
        'date_end':{'F': ['date'], 'V':['datepartial'], 'R':['yes']},
        'mutant_num':{'F': ['nospace'],'P': [ r'^\d+$' ],   'R':['yes']},
        'total_num':{'F': ['nospace'],'P': [ r'^\d+$' ],   'R':['yes']}
    }
}

parser = argparse.ArgumentParser(
    prog="GEOFF_tools",
    description="Utilities for GEOFF",
    formatter_class=argparse.ArgumentDefaultsHelpFormatter
)

subparsers = parser.add_subparsers(dest="command", required=True , help="Commands" )
subp_eef = subparsers.add_parser("excel_extract",  help="convert excel to 3 tsvs " )
subp_eef.add_argument ("--excel", help='excel file to create study_uid main and sites and mutprev', required=True )
subp_eef.add_argument ("--yaml", help='file of  preferences and parameters', required=True )

def excel_extract ( args ):
    #open     excel study data sheet
    #known issues -- getting .0 on interger values coming from parsing excel. pmid etc
    wb = openpyxl.load_workbook(args.excel)
    wb.sheetnames
    #open the yaml file
    with open (args.yaml, 'r') as f:
        settings=yaml.safe_load(f)
    #process the the excel file
    sheet_name=settings['excel']['sheet']['study']['name']
    if sheet_name not in wb.sheetnames:
        print(f"Sheet named '{sheet_name}' not found in the workbook.")
        sys.exit(1)
    ### (1) Study Overview sheet ###
    print (f"Extracting Study Overview Descriptions ({sheet_name})...")
    sheet = wb[sheet_name]
    study_column_names=['FIELDS','DATA']
    header_names, header_columns=find_headers_first_row(sheet,study_column_names)
    data =read_xls_sheet_columns(sheet, header_columns)
    print (f"...{header_names}  headers found in {header_columns} columns")
    data=read_xls_sheet_columns(sheet, header_columns)
    print (header_columns)
    data=data_delete_row_if_empty_cell(data,0)
    ### create a studydictonary of parameters to use down stream ###
    studydict= {row[0]: row[1] for row in data}
    studydict['study_uid']=studydict['study_uid'].lower()
    studydict['version']=studydict['version'].lower()
    study_uid_ver= studydict['study_uid']+"_"+ studydict['version']
    ##### die if the file names are not correct
    cwd=os.getcwd()
    if not study_uid_ver in cwd:
        log_it (f"!!!FATAL!!!: Current directory {cwd} is not named {study_uid_ver}. Please rename before proceeding")
        sys.exit(1)
    if not study_uid_ver in args.excel:
        log_it (f"!!!FATAL!!!: Current xls {cwd} does not start with {study_uid_ver}. Please rename before proceeding")
        sys.exit(1)
    ### get validation and formating data
    validation_list_study=validation_list_get(sheet)
    ### check for output directory
    if not os.path.exists( settings['subdir_extract'] ):
	    os.makedirs(settings['subdir_extract'] )
    save_data_tsv(data, settings['subdir_extract']+"/"+ study_uid_ver + para["file_suffix_study"]+".tsv",study_column_names)
    #extract_site_data
    validation_list_site=None
    if studydict['site_level_overview_complete']  == 'yes':
        sheet_name=settings['excel']['sheet']['site']['name']
        print (f"Extracting Site  Descriptions ({sheet_name})..")
        if sheet_name not in wb.sheetnames:
            print(f"Site Sheet named '{sheet_name}' not found in the workbook with sheets {wb.sheetnames}.")
            sys.exit(1)
        sheet = wb[sheet_name]
        prev_header_pattern= re.compile(r'SIT\w+')
        newheader_names, newheader_columns = find_headers_first_row (sheet,['FIELDS'], prev_header_pattern )
        print (f"...FIELDS headers found in {newheader_columns} columns...")
        if len(newheader_names) < 2:
            print (f'!!!!!!WARNING May be missing either FIELDS header or placeholding SURVEY')
        data=read_xls_sheet_columns(sheet, newheader_columns)
        data=data_delete_row_if_empty_cell(data,0)
        data=data_transpose(data,newheader_names)
        validation_list_site=validation_list_get(sheet)
        
        save_data_tsv(data, settings['subdir_extract']+"/"+ study_uid_ver+ para['file_suffix_site'] +".tsv",None)

    ### process site prevalences file
    prev_format=studydict['mutation_prev_site_complete']
    if  prev_format  in ('long', 'wide'):
        sheet_name=settings['excel']['sheet']['prev']['name']+"_"+prev_format
        print (f"Extracting Mutation Site Prevalences ({sheet_name})...")
        if sheet_name not in wb.sheetnames:
            print(f" Prev(alence) Sheet named '{sheet_name}' not found in the workbook.")
            sys.exit(1)
        sheet = wb[sheet_name]
        required_header_fields=settings['excel']['sheet']['prev'][prev_format]['fields']
        prev_header_pattern= re.compile(r'SIT|SURVE|\w+') #for wide
        if prev_format== 'long':
            prev_header_pattern = None
        newheader_names, newheader_columns = find_headers_first_row (sheet, required_header_fields, prev_header_pattern)
        print ( f"...{newheader_names} headers found in {newheader_columns} columns...")
        data=read_xls_sheet_columns(sheet, newheader_columns)
        data=data_delete_row_if_empty_cell(data,0)
        if prev_format=='wide': #transpose if wide format
            if  newheader_names[1] != 'date_start':  #then presume transposed by user with sites as rows aand mutations as columns
                print ("...transposing wide format")
                #need to transpose for easier processing later
                data=data_transpose(data,newheader_names,addheader=False)
                newheader_names=None
        save_data_tsv(data, settings['subdir_extract']+"/"+ study_uid_ver+ para['file_suffix_prev'] +"_"+ prev_format+".tsv",newheader_names)

    ### output validation formating as tsv (sheet\tfield\tparameter)
    if ( validation_list_site and  validation_list_study):
        validation_list_combined=  [ ]
        for row in validation_list_study: validation_list_combined.append(["study",row[1],row[0]] )
        for row in validation_list_site:  validation_list_combined.append(['site',row[1],row[0]])
        save_data_tsv(validation_list_combined, settings['subdir_extract']+"/"+ study_uid_ver+ para['file_suffix_validate'] + ".tsv", ["SHEET","KEY","VALUE"])

        if validation_list_study[len(validation_list_study)-1][1] != validation_list_site[len(validation_list_site)-1][1]:
            print (f"!!!!!!WARNING  the validation versions do not match--check for corruption")
    else:
        print (f"WARNING!!!!! No validation file created validation intructions data completely missing in study and/or site sheet. This error commonly occurs if the mutation_prev_site_complete in the study_overview tab on Google sheets is unchecked. Ensure it is set to the proper format (wide or long).")
    cwd=os.getcwd()

subp_tvf= subparsers.add_parser ("tsv_validate", help ="validate formats and check required data (coding in progress)",  )
subp_tvf.add_argument ("--study_tsv", help='tsv file with study info' , required=True )
subp_tvf.add_argument ("--validation_tsv", help='replace or add validation tsv (highest precendence is last)' )
subp_tvf.add_argument ("--yaml", help='file of  preferences and parameters', required=True )

def tsv_validate (args):
    log_it (str(args))
    log_it("Loading YAML...")
    with open (args.yaml, 'r') as f:
        settings=yaml.safe_load(f)
    basename=args.study_tsv.replace( (para["file_suffix_study"]+ ".tsv"), "")
    ###validation data
    vdict=None
    log_it("Loading validation parameters...")
    if args.validation_tsv:
        vdict=tsv_load_validate (args.validation_tsv)
    else:
        vdict=tsv_load_validate (basename+ para['file_suffix_validate']+".tsv")
    
    ###study data

    log_it("Loading Study data...")
    study_data=tsv_load_datadict (basename+ para['file_suffix_study']+".tsv")#
    log_it("Validating Study data...")
    for row in study_data:
        cleanfield, is_valid, vf_errors = text_main_format_validate (row['DATA'],vdict['study'][row['FIELDS']])
        #print (row['FIELDS'],"|", row['DATA'], "|",cleanfield, is_valid, vf_errors)
        if not is_valid:
            log_it (f"...!!!{row['FIELDS']} bad VorF: {row['DATA']} {vf_errors}" )
        row['DATA']=cleanfield
    study_dict={}
    for row in study_data:
        study_dict[str(row['FIELDS'])]=str(row['DATA'])
    log_it (f"...validated study data saving to validated.tsv")
    save_data_tsv(study_data, basename + para["file_suffix_study"]+"_validated.tsv",['FIELDS','DATA'])

    if study_dict['site_level_overview_complete']=='yes'  :  ### add args.force_site  yes
        log_it ("Loading and Validating Site data (wide tsv)...")
        site_data=tsv_load_datadict (basename+ para['file_suffix_site']+".tsv")
        log_it ("...formating and validating data.")
        for row in site_data:
            for k in row.keys():
                cleanfield, is_valid, vf_errors = text_main_format_validate (row[k], vdict['site'][k] )
                if not is_valid:
                    log_it ('...!!!' + k + cleanfield + str(is_valid) + "ERRORS: " +  vf_errors)
                row[k]=cleanfield
        all_sites=sorted([ row['site_uid'] for row in site_data])
        dedup_all_sites=sorted(set(all_sites))
        log_it ( "...list of sites"+ str(all_sites) )
        if len(dedup_all_sites) < len(all_sites):
            log_it('!!!FATAL!!! ERROR REDUNDANT site_uid')
            log_it (str(len(all_sites)-len(dedup_all_sites)) + " site(s) are redundant")
            sys.exit (1)

        ###Write a file out here ####
        log_it( f"...saving data validated.tsv")
        save_data_tsv(site_data, basename + para["file_suffix_site"]+"_validated.tsv",site_data[0].keys())
    log_it ("Loading and Validating Prev(alence) data...")
    psite_dict= defaultdict(int) #hash to store prevalence sites
    survey_dict=defaultdict(int)  #combinations of the survey data
    mutations=defaultdict(int)
    long_data=[]  #we will move to this type
    if study_dict['mutation_prev_site_complete'] == 'wide':
        log_it ("...loading wide format.")
        prev_data=tsv_load_datadict (basename+ para['file_suffix_prev']+"_"+study_dict['mutation_prev_site_complete']+".tsv")
        headers=list(prev_data[0].keys())
        headers = { headers[i].strip(): headers[i].strip() for i in range(len(headers)) }
        fatal_error=False
        for h in headers.keys():
            is_valid=True
            if h  in settings['excel']['sheet']['prev']['wide']['fields'] :
                good=is_valid#print (h , "normal field")
            elif re.search(r'^[A-Za-z0-9_-]+:[0-9_-]+:[A-Za-z0-9*_|/]+$', h ):
                print (h)
                is_valid, logerror,newlongtext=long_form_validate(h,type='protein')
                if is_valid:
                    headers[h]=newlongtext
            else:
                log_it ("*!*!*!*! FATAL ERROR BAD PREV HEADER")
                log_it (f"!!!{h} is not a valid header or longform feature:positions:variants")
                fatal_error=True
        if fatal_error:
            log_it (" FATAL: Fix please: e.g   CRT:76:T or crt:73_74_75_76:CVIET")
            sys.exit(1)
        row_count=1
        long_row_headers= settings['excel']['sheet']['prev']['wide']['fields']
        fatal_rows=0
        fatal_errors=0
        for row in prev_data:
            row_count+=1
            long_row_start= [ row[h] for h in settings['excel']['sheet']['prev']['wide']['fields'] ]
            logerror=''
            ##validate and refomrat
            long_row_start, all_valid, fatal_count ,survey_error =clean_survey_fields(long_row_start,
                settings['excel']['sheet']['prev']['wide']['fields'], para["prev_field_format"])
            if len(long_row_start) !=  len(settings['excel']['sheet']['prev']['wide']['fields']):
                log_it ("PARSING FAILED")
                sys.exit(1)
            if not all_valid:
                if fatal_count:
                    logerror+= f"!!!FATAL survey fields {survey_error}; "
                    fatal_errors+=fatal_count
                else:
                    logerror+=f"!!!BAD survey fields {survey_error}; "
            ###check if name is unique
            survey_dict[":".join(long_row_start)]+=1
            psite_dict[long_row_start[0]]+=1
            for k,v in row.items():
                if not k in settings['excel']['sheet']['prev']['long']['fields']:
                    #for each mutation spit out a row (it will be long ;-)
                    if v =="": continue
                    newrow = long_row_start.copy()
                    newrow.append( headers[k] )  ## this is the clean name
                    fraction, newerror = field_fraction_split_validate(v)
          
                    if fraction!=None:
                        newrow.extend(fraction)
                        mutations[headers[k]]+=1
                    else:
                        newrow.extend(["BAD","BAD"])
                        logerror += f"; {newerror}"
                        fatal_errors+=1
                    long_data.append(newrow)
            if logerror !="":
                log_it ( f"...!!!Row{row_count} ERROR:{logerror}")
        log_it (f"... {len(long_data)} long rows from conversion of wide format")
        if fatal_errors>0:
            log_it( f"!*!*!FATAL ERRORS ABOVE : {fatal_errors}  BAD ROW(s) ")
            sys.exit(1)
        
    if study_dict['mutation_prev_site_complete'] == 'long':
        long_headers=settings['excel']['sheet']['prev']['long']['fields']
        log_it ("...loading long format.")
        prev_data=tsv_load_datadict (basename+ para['file_suffix_prev']+"_"+study_dict['mutation_prev_site_complete']+".tsv")
        log_it ("...checking header names")
        for h in   prev_data[1].keys():
            if not h in long_headers:
                log_it(  "!!!FATAL!!!:  {h} not a long_header -- bad column names")
                sys.exit()
        row_count=0
        fatal_long=0
        for row in prev_data:
            logerror=''
            fatal=0
            row_count+=1
            for k in row.keys():
                if k in para['prev_field_format']:
                    cleanfield, is_valid, errortext =text_main_format_validate(row[k], para['prev_field_format'][k])
                    #print (k, cleanfield, is_valid, errortext )
                    if not is_valid:
                        logerror += f"{errortext}; "
                    row[k]=cleanfield
                elif  k == 'gene_mutation':
                    is_valid, lferror,newlongtext=long_form_validate(row[k],type='protein')
                    if is_valid :
                        row[k]=newlongtext
                    else:
                        logerror += f"...!FATAL! longformat bad {lferror};"
                        fatal_long+=1
                else:
                    log_it(  "...!FATAL!:  {k} not a known entry parsing through prev_data")
                    fatal_long+=1
            
            #ow['mutant_num']=re.sub(r'\.0',"",row['mutant_num'])
            #row['total_num']=re.sub(r'\.0',"",row['total_num'])
            try:
                #Checks that values entered are sample counts -- either intger as expected, or a integer-like float
                if (float(row['mutant_num']).is_integer() == False) or (float(row['total_num']).is_integer() == False):
                    logerror+= f"{row['mutant_num']} or {row['total_num']} non-integer-like. Check that long-style data entry sheet contains integer sample counts and totals."

                if  int(float(row['mutant_num'])) > int(float(row['total_num'])):
                    logerror+= f"more mutants {row['mutant_num']} than total {row['total_num']}; "
            except:
                logerror+= f"Bad intergers or math mut# {row['mutant_num']} row#  {row['total_num']}"
                fatal_long+=1
            if logerror !=None and logerror != "":
                log_it (f"...ROW {row_count} {logerror}")
            psite_dict[row['site_uid']]+=1
            survey_dict[":".join([row['site_uid'],row['substudy'],row['date_start'],row['date_end'] ])]+=1
            mutations[row['gene_mutation']]+=1
        long_data=prev_data
        if fatal_long:
            log_it (f"!!!FATAL ERRORS: {fatal_long} found that need to be fixed")
            sys.exit(1) 

    if len (long_data)>=1:
        log_it ("SITES FROM PREVALENCE:" + str( dict(sorted(psite_dict.items(), key = lambda item: item[1])) ))
        for s in psite_dict.keys():
            if not s in dedup_all_sites:
                log_it (f"!!!FATAL ERROR!!!: {s} is not in Sites:{dedup_all_sites}")
                exit()
        log_it ("SURVEYS:" + str( dict(sorted(survey_dict.items(), key = lambda item: item[1]))))
        log_it ("MUTATIONS"+ str( dict(sorted(mutations.items(), key = lambda item: item[1]))))
        save_data_tsv(long_data, basename + para["file_suffix_prev"]+"_LONG_validated.tsv"
             ,settings['excel']['sheet']['prev']['long']['fields'])

def clean_survey_fields(survey_fields, survey_headers, format_fields):
    logerror=''           ### normally site_uid, substuy, date_start, date_end
    all_valid=True
    fatal_count=1
    new_fields=[]
    for  k,v in  zip(survey_headers,survey_fields  ):
        cleanfield , is_valid, errortext =text_main_format_validate(v, format_fields[k])
        if not is_valid:
           logerror+=f"{errortext}; "
           all_valid=False
           if "R::" in errortext:
               fatal_count+=1
        new_fields.append(cleanfield)
    return new_fields,  all_valid, fatal_count, logerror

###tsv loaders
def tsv_load_validate (file):
    vdict={'study':{},'site': {} }
    validate_parse_error_count=0
    with open (file, 'r') as f:
        reader = csv.DictReader(f, delimiter = '\t')
        for row in reader:
            if not len (row) ==3 :
                log_it("WARNING" )
            row['KEY']= re.sub(r'\s+', ' ', row['KEY'].strip() )
            row['VALUE']= re.sub(r'\s+', ' ', row['VALUE'].strip())
            vdict[row['SHEET']][row['KEY']] = rdict=defaultdict(list)

            if not row['VALUE']: continue
            if row['KEY']=='VERSION':
                vdict[row['SHEET']][row['KEY']][row['VALUE']]==row['VALUE']
                continue
            for typetext in  row['VALUE'].split(";"):
                try:
                    type, text = typetext.split("::")
                    vdict[row['SHEET']][row['KEY']][type].append(text)
                except:
                    print( f"!!!!!!{row['VALUE']}  bad format in {row['SHEET']} {row['KEY']} ILL FORMED split on '::'")
                    validate_parse_error_count= validate_parse_error_count+1
    if validate_parse_error_count :
        print (f"!!!!{validate_parse_error_count} PARSING VALIDATION FILE HAD MULTIPLE ERRORS")
    return vdict

def tsv_load_datadict(file):
    data=[]
    with open (file, 'r') as f:
        reader = csv.DictReader(f, delimiter = '\t')
        for row in reader:
            data.append(row)
    return data

###this splits NN/DDD fractions in wide table and validates
def field_fraction_split_validate(field):
    ### look for NN/NNNN None or [NN,DDD] text array
    
    matched=field.strip().split("/")

    if len(matched)!=2:
        return None, f"ERROR {matched} is not two items {field}"
        
    matched[0]=re.sub(r'\.0',"",matched[0])
    matched[1]=re.sub(r'\.0',"",matched[1])

    if int(matched[0]) > int (matched[1]):
        return None, f"ERROR {field} {matched[0]} mutant > {matched[1]} total "
    
    return matched,None

#validating specific to the long form variant/haplotype foramt
def long_form_validate(longformat, type=None, valid_genes=None):
    logerror=None
    is_valid=True
    if not re.search( r'^[A-Za-z0-9_-]+:[0-9_-]+:[A-Za-z0-9*_|/]+$', longformat ):
        is_valid=False
        logerror=f"Does not match main pattern ({longformat}) for region:positions:variants"
    else:
        lf_match=re.match(r'^([A-Za-z0-9_-]+):([0-9_-]+):([A-Za-z0-9*_|/]+)$', longformat)
        if  len(lf_match.groups())==3:
            site,positions,variants=lf_match.groups()
            if  variants in ['*',"ALL_V","ALL_C","ALL_S","ALL_ALL","ALL_VC"]   and re.search(r'^\d+-\d+$',positions):
                is_valid=True #special star foramt
            elif "_" in positions and "_" in variants:
                if len (positions.split("_")) != len (variants.split("_")):
                    is_valid=False
                    logerror=f"Does not split properly _ ({longformat}) for equal # postions and # variants"
            elif "_" in positions: #then  we have issues of
                if len (positions.split("_")) !=  len (list(variants )):
                    logerror=f"Does not split properly ({longformat}) for equal # multiple positions and nonseparted _ single AA variants"
        else :
            logerror=f"Does not split properly on colons ({longformat}) for region:positions:variants"
        if is_valid:
            longformat=":".join([site,positions,variants])
            if type=='protein': longformat=longformat.upper()
    return is_valid, logerror,longformat

#### the following 5 routines do general formating and validating
def text_main_format_validate(field_text, commands_dict):
    check_order=[
        'F'  , #do a formatting first
        'S', # do substitutions next
        ### validation last
        'P' ,#: pattern (validation complains if bad)
        'V', #: other validations not handled by a simple pattern
        'R', #r#equired (if empty complain bitterly)
    ]
    vf_errors=''
    is_valid=True
    newvalue=field_text
    for letter in sorted (check_order):
        if letter in commands_dict:
            for command in commands_dict[letter]:
                errortext=None
                is_valid_still=True
                if letter == "F":
                    newvalue,errortext=text_Format(newvalue, command)
                elif letter =="S":
                    newvalue,errortext=text_Substitution (newvalue,command)
                elif letter =="P":
                    is_valid_still, errortext=text_validate_Pattern (newvalue,command)
                elif letter =="V":
                    is_valid_still, errortext= text_validate_Validate (newvalue,command)
                elif letter =="R":
                    if newvalue=="":
                       errortext = f"Empty {letter}::{command}"
                else:
                    errortext = f"{letter} invalid {letter}::{command}"
                if errortext != None and errortext !='':
                    vf_errors +=  str(errortext) + ";"
                if not is_valid_still:
                    is_valid=False
    return newvalue, is_valid, vf_errors###main command

def text_Format(ftext, command):
        ferror=None
        if  command=='lowercase':
            ftext= ftext.lower()
        elif command=='uppercase':
            ftext= ftext.upper()
        elif command=='titlecase':
            ftext= ftext.title()
        elif command=='nospace':
            ftext=re.sub(r'\s+',"",ftext)
        elif command=='date':
            ftext=re.sub(r'_',"-",ftext)
            ftext=re.sub(r'\s+',"",ftext)
            ftext=re.sub(r'\.0',"",ftext)
        elif command=='onespace':  #removes
            ftext=re.sub(r'\s+'," ",ftext.strip())
            ftext=re.sub(r'\s+,',",",ftext)
            ftext=re.sub(r',\s+',",",ftext)
        else:
            ferror=f"{command} unknown F::{command} VALID:lowercase uppercase titlecase nospace onepsace date!;"
        return ftext, ferror  #used by main

def text_Substitution (ftext,findreplace):
    serror=None
    try:
        find,replace=findreplace.split(":")
        ftext=re.sub(fr"{find}", replace, ftext)
    except:
        serror= f"S::{findreplace} illformed? S::findpattern:replacetext"
    return ftext, serror#used by main

def text_validate_Pattern (ftext, pattern):
    ###only validates no
    is_valid=False
    perror=None
    if ftext =='' or re.search(fr"{pattern}",  ftext):
        is_valid=True
    else:
        perror=f"'{ftext}' failed 'P::{pattern} pattern"
    #print (is_valid,perror)
    return is_valid, perror#used by main

def text_validate_Validate(ftext, command):
    is_valid=False
    verror=None
    if command=='datefull': #YYYY-MM-DD
        if re.search(r'^[1-2]\d\d\d-[0-1]\d-[0-3]\d$', ftext):
            is_valid=True
    elif command=='datepartial':
        if re.search(r'^[1-2]\d\d\d-[0-1]\d-[0-3]\d$|^[1-2]\d\d\d-[0-1]\d$|^[1-2]\d\d\d$', ftext):
            is_valid=True
    elif command=='word':
         if re.search(r'^[\w-]+$', ftext):
            is_valid=True
    elif command=='float':
         if re.search(r'^[+-]?[0-9]+.?[0-9]?$',ftext):
             is_valid=True
    elif command=='iso3c':
        if ftext in para['iso3c_africa'] :
                is_valid=True
    elif command=='filepath':
        if  os.path.exists (ftext): 
            is_valid=True
    elif re.search(r'^min:|^max:',command) :
        try:
            subcommand,value= command.split(":")
            if subcommand == 'min':
                if ftext == '' or float(ftext) >= float(value): is_valid=True
                else: verror = f"{ftext}  is below threshold {command}"
            elif subcommand =='max':
                if ftext == '' or float(ftext) <= float(value): is_valid=True
                else: verror = f"{ftext}  is above threshold {command}"
            else:
                verror = f"{command} code error in V--should never get here"
        except:
            verror = f"{command} on {ftext} failed parsing or math calcs on numbers"
    else:
        verror=f"{command} unknown V::{command} to validate!;"
        #isprint (command)
    if is_valid==False and verror == None:
        verror =f"'{ftext}' failed 'V::{command}'"
    return is_valid,verror#used by main

######################
###shared functions

def validation_list_get ( sheet):
    #load the column validate_format and fields
    header_list=[]
    print ("...hunting validation")
    re_pattern=re.compile(r'VALIDATE_FORMAT')
    data=[]
    header_names, header_columns=find_headers_first_row(sheet,['FIELDS'],re_pattern)
    if len(header_names)!=2:
        log_it (f"!!!!!!WARNINGno validation formating found --need VALIDATE_FORMAT and FIELDS -- just {header_names}.")
        log_it (f"!!!!!!For proper validaton in tsv_validate you will need external validate tsv")
        return data
    data=read_xls_sheet_columns(sheet, header_columns)
    data=data_delete_row_if_empty_cell(data,1)
    ver_pattern=re.compile(r'VERSION:(V\d+[.]\d\d\d\d_\d\d_\d\d)')
    matched=ver_pattern.findall(header_names[0])
    if matched:
        data.append([matched[0],'VERSION'])
        log_it (f"......Version: {matched[0]} of validation")
    else :
        log_it (f"!!!!!!!!!WARNING!  validation not versioned")
    log_it  ( f"......{len(data)} fields of formatting")
    if (len (data) <5):
        log_it (f"!!!!!!!!!WARNING! your validation may be truncated")
    return data #this pulls data from excel columns in study and site sheets

def data_transpose(data, headers,addheader=False):
    transposed_data=[] #reuse data to transpose
    for c in range(len(headers)):
        newrow=[]
        if addheader:  newrow.append(headers[c])
        for row in data: newrow.append (row[c])
        transposed_data.append(newrow)
    return transposed_data  #can flip a prevalence wide format

def data_delete_row_if_empty_cell(data, column_number):
    cleaned_data = [ ] #remove  empty first cells.
    for row in data:
        if   row[column_number].strip() == '':
            continue
        cleaned_data.append(row)
    return cleaned_data  #based on a column being empty delete entire row

def find_headers_first_row ( sheet, exact_header_list, re_pattern=None):
    headers = []
    #print ([cell.value for cell in sheet[1]])
    for cell in sheet[1]:
       if cell.value is None:
          headers.append("")
       else :
          headers.append(str(cell.value))
    headers = [ re.sub(r'\s+', ' ', value.strip() ) for value in headers]
    newheader_names=[]
    newheader_columns=[]
    for i, v in enumerate(headers):
        if v=='':
            continue
        good=False
        if re_pattern is None:
            if  v  in exact_header_list:
                good=True
        elif v  in exact_header_list or re_pattern.match(v):
            good=True

        if good:
            newheader_names.append(v)
            newheader_columns.append(i)
    return (newheader_names, newheader_columns)  #finds specific columns that can then pull subseuqent rows

def read_xls_sheet_columns(sheet, column_numbers):
    # Collect the data from the specified columns
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        newrow = []
        for i in column_numbers:
            if row[i] is None:
                newrow.append( '')
            else:
                newrow.append(  str(row[i])   )
        newrow = [ re.sub(r'\s+', ' ', value.strip() )  for value in newrow]
        data.append(newrow)
    return data  #read specific headers from excel

def save_data_tsv ( data, output_tsv_file_path, header_names) :# Save  TSV file
    #this reorganize
    with open(output_tsv_file_path, 'w', newline='', encoding='utf-8') as tsv_file:
        if header_names:
            tsv_file.write('\t'.join(header_names) + '\n')
        # Write data rows and remove empty rows
        for row in data:
            if isinstance(row, dict):
                tsv_file.write ('\t'.join([row[k] for k in row.keys()]) +"\n")
            else:
                tsv_file.write('\t'.join([str(cell) if cell is not None else '' for cell in row]) + '\n')
    print(f"...{output_tsv_file_path} saved containing {len(data)} rows data")

def log_it (message):
    print(message)

def main (args):
  if args.command == 'excel_extract':
    excel_extract(args)
  elif args.command == 'tsv_validate':
    tsv_validate(args)
  elif args.command == 'siteprev_hack':
    print ( "This command is not implmented yet!")
    sys.exit()
  else:
    print ("command:(", args.command, ") does not exist!")
    sys.exit()

if __name__ == '__main__':
  #figure out how to make arguments from the main transfer to the subs
  #add a --quiet command

  parser._positionals.title = 'command'  #removes Positional in help""
  parser.add_argument( "--version",  action="version",
    version="%(prog)s " + __version__
  )

  #to be run by others
  subp_shf= subparsers.add_parser ("siteprev_hack", help ="generate a _tv site long prevalence file for hackathonV (coming soon)" )
  subp_shf.add_argument ("--study_tsv", help='tsv file from study overview' , required=True )
  subp_shf.add_argument ("--yaml", help='file of  preferences and parameters', required=True )
  subp_shf.add_argument ("--study_only", help='only validate study file (skip site and prev)', action="store_true", default=False )

 # subp_af= subparsers.add_parser ('all', help ="generate a  site long prevalence file for hackathonV" )
 # subp_af.add_argument ('--excel', help='run  extraction, validation, and creation of siteprev  in one fell swoop', required=True  )

  args=parser.parse_args()
  main(args)
